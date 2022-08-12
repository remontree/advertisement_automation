from selenium import webdriver
import time
import pyautogui
import pyperclip
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
import openpyxl
from selenium.webdriver.support.select import Select
import pandas as pd
import datetime
import autoit

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument('--disable-dev-shm-usage')

save_path = "정보.xlsx"
save_path2 = "소스만.xlsx"
# 기존 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True)
wa = openpyxl.load_workbook(save_path2, data_only=True)
sh = pd.read_excel("정보.xlsx")
#시트선택
ws = wb.active
ws2 = wa.active
chromedriver_autoinstaller.install()

def upload(a, wb=wb,wa=wa):
    url = "https://www.ppomppu.co.kr/zboard/login.php?r_url=http%3A%2F%2Fwww.ppomppu.co.kr%2Fzboard%2Fzboard.php%3Fid%3Dpmarket3"
    browser=webdriver.Chrome()
    browser.implicitly_wait(100) # 로딩을 기다릴것 최대 100초
    browser.maximize_window() # 화면 최대화할것
    browser.get(url)
    # 로그인
    id = browser.find_element(by=By.XPATH, value="/html/body/div/form/ul/li[1]/input[1]")
    id.click()
    id.click()
    id.send_keys(ws.cell(row=a,column=1).value)

    # 패스워드

    pw = browser.find_element(by=By.XPATH, value="/html/body/div/form/ul/li[1]/input[2]")
    pw.click()
    pw.click()
    pw.send_keys(ws.cell(row=a,column=2).value)

    # 로그인 버튼 누르기
    log_btn = browser.find_element(by=By.XPATH, value="/html/body/div/form/ul/a")
    log_btn.click()
    time.sleep(2.5)

    # 글쓰기
    write_add = browser.find_element(by=By.XPATH, value="/html/body/div[6]/div[2]/div[5]/div[1]/table[4]/tbody/tr[2]/td[2]/nobr[2]/a/font")
    write_add.click()
    time.sleep(2.5)

    # 제목
    title = browser.find_element(by=By.XPATH, value="/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[2]/tbody/tr[3]/td/input")
    title.send_keys(ws.cell(row=a,column=3).value)
    time.sleep(2.5)

    #링크
    link = browser.find_element(by=By.XPATH, value="/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[2]/tbody/tr[4]/td/input")
    link.send_keys(ws.cell(row=a,column=4).value)
    time.sleep(2.5)

    # 내용
    edit = browser.find_element(by=By.XPATH, value="/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[2]/tbody/tr[6]/td/div[3]/div[6]/div[2]")
    edit.click()
    pyperclip.copy(ws2.cell(row=2,column=1).value)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(1)
    edit = browser.find_element(by=By.XPATH, value="/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[2]/tbody/tr[6]/td/div[3]/div[6]/div[1]")
    edit.click()
    time.sleep(2.5)

    #분류
    select_element = browser.find_element(by=By.XPATH,value='/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[2]/tbody/tr[2]/td/span/select')
    select_object = Select(select_element)
    select_object.select_by_visible_text('인터넷')

    #이미지 업로드
    ssumnail = browser.find_element(by=By.XPATH,value='/html/body/div[6]/div[2]/div[5]/div/form/table/tbody/tr/td/table[3]/tbody/tr[5]/td/div[2]/span')
    ssumnail.click()
    time.sleep(5)

    #Basic Window info 값 handle 변수에 저장
    handle = "[CLASS:#32770; TITLE:열기]"

    #이름이 '열기'인 창이 나올 때까지 3초간 대기
    autoit.win_wait_active("열기", 3)

    img_path = '"C:\\Users\\remon\\OneDrive\\바탕 화면\\제품사진.jpg"'

    #사진 클릭시 나오는 윈도우 창에서 파일이름(N)에 이미지 경로값 전달
    autoit.control_send(handle, "Edit1", img_path)
    time.sleep(1)

    #사진 클릭시 나오는 윈도우 창에서 Button1 클릭
    autoit.control_click(handle, "Button1")
    time.sleep(1)

    #  이벤트 -> 파일이 다 합쳐진 상태로 왔을 때 처리

    # 종료
    #browser.close()
